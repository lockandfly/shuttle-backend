import csv
import io
from datetime import datetime
import openpyxl

from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.bookings.models_orm import Booking


class ImportService:

    @staticmethod
    def import_file(db: Session, file: UploadFile, portal: str):
        filename = file.filename.lower()

        # Carica contenuto
        if filename.endswith(".csv"):
            content = file.file.read().decode("utf-8")
            rows = list(csv.DictReader(io.StringIO(content)))
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file.file)
            ws = wb.active
            rows = list(ImportService._xlsx_to_dict(ws))
        else:
            raise HTTPException(status_code=400, detail="Formato file non supportato. Usa CSV o XLSX.")

        bookings = []

        for row in rows:
            try:
                booking = ImportService._parse_parkos_row(row, portal)
                db.add(booking)
                bookings.append(booking)
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Errore parsing riga: {row} → {str(e)}"
                )

        db.commit()
        return bookings

    @staticmethod
    def _xlsx_to_dict(ws):
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(headers, row))

    @staticmethod
    def _parse_parkos_row(row: dict, portal: str) -> Booking:

        def parse_date(value):
            if not value:
                return None
            try:
                return datetime.fromisoformat(value.replace(" ", "T"))
            except:
                return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")

        return Booking(
            portal=portal,

            code=row.get("code"),
            customer_name=row.get("name"),
            customer_email=row.get("email"),
            phone=row.get("phonenumber"),
            license_plate=row.get("car_sign"),

            arrival=parse_date(row.get("arrival")),
            departure=parse_date(row.get("departure")),

            price=float(row.get("price")) if row.get("price") else None,

            payment_complete=row.get("paymentComplete") in ["sì", "yes", "true", "1"],
            external_id=row.get("externalID"),
            online_payment=row.get("onlinePayment") in ["sì", "yes", "true", "1"],
            payment_option=row.get("paymentOption"),

            cancel_date=parse_date(row.get("cancelDate")),
            cancel_reason=row.get("cancelReason"),

            passengers=int(row.get("numberOfPassengers")) if row.get("numberOfPassengers") else None,
            days=int(row.get("numberOfCalendarDays")) if row.get("numberOfCalendarDays") else None,

            notes=None
        )
