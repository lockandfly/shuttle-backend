from fastapi import APIRouter, UploadFile, File, HTTPException
from typing import List
import openpyxl
import csv
import io
from datetime import datetime, timedelta

from app.bookings.portal_enum import Portal
from app.bookings.importer.myparking import MyParkingImporter
from app.bookings.importer.parkos import ParkosImporter
from app.bookings.importer.parkingmycar import ParkingMyCarImporter
from app.bookings.models import BookingRead

from app.database import SessionLocal
from app.bookings.models_orm import Booking, Shuttle, ShuttleAssignment

router = APIRouter()

IMPORTERS = {
    Portal.myparking: MyParkingImporter(),
    Portal.parkos: ParkosImporter(),
    Portal.parkingmycar: ParkingMyCarImporter(),
}

# -------------------------
# FUNZIONI DI SUPPORTO
# -------------------------

def detect_headers(file: UploadFile):
    filename = file.filename.lower()

    if filename.endswith(".csv"):
        content = file.file.read().decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(content))
        headers = next(reader)
        file.file.seek(0)
        return [h.strip() for h in headers]

    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        headers = [str(h).strip() for h in ws[1]]
        file.file.seek(0)
        return headers

    raise HTTPException(status_code=400, detail="Formato file non supportato")


def auto_detect_portal(file: UploadFile):
    filename = file.filename.lower()

    if filename.endswith(".csv"):
        content = file.file.read().decode("utf-8", errors="ignore")
        headers = next(csv.reader(io.StringIO(content)))
        file.file.seek(0)

        h = [h.strip().lower() for h in headers]

        if "code" in h and "email" in h and "car_sign" in h:
            return Portal.parkos

        if "prenotazione" in h or "codice prenotazione" in h:
            return Portal.myparking

        return Portal.parkingmycar

    if filename.endswith(".xlsx"):
        return Portal.myparking

    raise HTTPException(status_code=400, detail="Formato file non supportato")

# -------------------------
# ENDPOINTS BASE IMPORT
# -------------------------

@router.post("/preview")
async def preview_file(file: UploadFile = File(...)):
    headers = detect_headers(file)

    filename = file.filename.lower()
    rows_preview = []

    if filename.endswith(".csv"):
        content = file.file.read().decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(content))
        next(reader)
        for i, row in enumerate(reader):
            if i >= 3:
                break
            rows_preview.append(row)
        file.file.seek(0)

    elif filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if i > 3:
                break
            rows_preview.append(list(row))
        file.file.seek(0)

    return {
        "filename": file.filename,
        "headers": headers,
        "preview_rows": rows_preview,
    }


@router.post("/validate")
async def validate_file(
    portal: Portal,
    file: UploadFile = File(...)
):
    importer = IMPORTERS[portal]

    try:
        bookings: List[BookingRead] = importer.parse(file.file, file.filename)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    return {
        "valid": True,
        "count": len(bookings),
        "sample": bookings[:3],
    }


@router.post("/import")
async def import_bookings(
    portal: Portal,
    file: UploadFile = File(...)
):
    importer = IMPORTERS[portal]

    try:
        bookings = importer.parse(file.file, file.filename)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    db = SessionLocal()

    for b in bookings:
        record = Booking(
            portal=b.portal,
            portal_reservation_id=b.portal_reservation_id,
            customer_name=b.customer_name,
            customer_email=b.customer_email,
            customer_phone=b.customer_phone,
            car_plate=b.car_plate,
            car_model=b.car_model,
            amount=b.amount,
            status=b.status,
            parking_area=b.parking_area,
            note=b.note,
            payment_method=b.payment_method,
            cancel_reason=b.cancel_reason,
            passenger_count=b.passenger_count,
            calendar_days=b.calendar_days,
            checkin=b.checkin,
            checkout=b.checkout,
            created_at=b.created_at,
            updated_at=b.updated_at,
        )
        db.add(record)

    db.commit()
    db.close()

    return {"imported": len(bookings)}


@router.get("/db/bookings")
def get_bookings():
    db = SessionLocal()
    data = db.query(Booking).all()
    db.close()
    return data

# -------------------------
# OCCUPAZIONE BASE + AVANZATA
# -------------------------

@router.get("/db/occupancy")
def get_occupancy():
    now = datetime.now()
    db = SessionLocal()

    active = db.query(Booking).filter(
        Booking.checkin <= now,
        Booking.checkout >= now
    ).count()

    db.close()

    return {
        "timestamp": now,
        "cars_in_parking": active
    }


@router.get("/db/occupancy/advanced")
def occupancy_advanced():
    now = datetime.now()
    today = now.date()

    db = SessionLocal()

    # 1. Auto presenti ora
    current = db.query(Booking).filter(
        Booking.checkin <= now,
        Booking.checkout >= now
    ).count()

    # 2. Occupazione per area
    areas = {}
    all_areas = db.query(Booking.parking_area).distinct().all()

    for (area,) in all_areas:
        if not area:
            continue
        count = db.query(Booking).filter(
            Booking.parking_area == area,
            Booking.checkin <= now,
            Booking.checkout >= now
        ).count()
        areas[area] = count

    # 3. Occupazione per giorno (oggi)
    start_day = datetime(today.year, today.month, today.day)
    end_day = start_day + timedelta(days=1)

    daily = db.query(Booking).filter(
        Booking.checkin <= end_day,
        Booking.checkout >= start_day
    ).count()

    # 4. Occupazione per fascia oraria (ogni 2 ore)
    hourly = []
    for h in range(0, 24, 2):
        start = datetime(today.year, today.month, today.day, h)
        end = start + timedelta(hours=2)

        count = db.query(Booking).filter(
            Booking.checkin <= end,
            Booking.checkout >= start
        ).count()

        hourly.append({
            "range": f"{h:02d}:00 - {h+2:02d}:00",
            "cars": count
        })

    db.close()

    return {
        "timestamp": now,
        "current": current,
        "by_area": areas,
        "today_total": daily,
        "hourly": hourly
    }
# -------------------------
# NAVETTE — CREAZIONE, LISTA, STATO
# -------------------------

@router.post("/shuttle/create")
def create_shuttle(name: str, capacity: int):
    db = SessionLocal()

    existing = db.query(Shuttle).filter(Shuttle.name == name).first()
    if existing:
        db.close()
        raise HTTPException(status_code=400, detail="Navetta già esistente")

    shuttle = Shuttle(name=name, capacity=capacity, status="idle")
    db.add(shuttle)
    db.commit()
    db.refresh(shuttle)
    db.close()

    return {"created": shuttle}


@router.get("/shuttle/list")
def list_shuttles():
    db = SessionLocal()
    shuttles = db.query(Shuttle).all()
    db.close()
    return shuttles


@router.get("/shuttle/status")
def shuttle_status():
    db = SessionLocal()
    shuttles = db.query(Shuttle).all()
    db.close()

    return {
        "status": "ok",
        "shuttles": shuttles
    }

# -------------------------
# NAVETTE — CODE ARRIVI E PARTENZE
# -------------------------

@router.get("/shuttle/queue/arrivals")
def shuttle_queue_arrivals():
    now = datetime.now()
    db = SessionLocal()

    upcoming = db.query(Booking).filter(
        Booking.checkin >= now
    ).order_by(Booking.checkin.asc()).limit(30).all()

    db.close()
    return upcoming


@router.get("/shuttle/queue/departures")
def shuttle_queue_departures():
    now = datetime.now()
    db = SessionLocal()

    leaving = db.query(Booking).filter(
        Booking.checkout <= now + timedelta(hours=2)
    ).order_by(Booking.checkout.asc()).limit(30).all()

    db.close()
    return leaving

# -------------------------
# NAVETTE — ASSEGNAZIONE AUTOMATICA
# -------------------------

@router.post("/shuttle/assign/next")
def assign_next_shuttle():
    now = datetime.now()
    db = SessionLocal()

    # 1. Trova la prossima prenotazione in arrivo
    next_booking = db.query(Booking).filter(
        Booking.checkin >= now
    ).order_by(Booking.checkin.asc()).first()

    if not next_booking:
        db.close()
        return {"assigned": False, "reason": "Nessuna prenotazione imminente"}

    # 2. Trova una navetta disponibile
    shuttle = db.query(Shuttle).filter(
        Shuttle.status == "idle"
    ).order_by(Shuttle.id.asc()).first()

    if not shuttle:
        db.close()
        return {"assigned": False, "reason": "Nessuna navetta disponibile"}

    # 3. Assegna la navetta
    assignment = ShuttleAssignment(
        shuttle_id=shuttle.id,
        booking_id=next_booking.id,
        timestamp=datetime.now()
    )

    shuttle.status = "loading"

    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    db.close()

    return {
        "assigned": True,
        "shuttle": shuttle.name,
        "booking_id": next_booking.id,
        "customer": next_booking.customer_name,
        "checkin": next_booking.checkin
    }
# -------------------------
# DASHBOARD OPERATIVA COMPLETA
# -------------------------

@router.get("/dashboard")
def dashboard():
    now = datetime.now()
    db = SessionLocal()

    # -------------------------
    # 1. ARRIVI (prossime 6 ore)
    # -------------------------
    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(hours=6)
    ).order_by(Booking.checkin.asc()).all()

    # -------------------------
    # 2. PARTENZE (prossime 6 ore)
    # -------------------------
    departures = db.query(Booking).filter(
        Booking.checkout >= now,
        Booking.checkout <= now + timedelta(hours=6)
    ).order_by(Booking.checkout.asc()).all()

    # -------------------------
    # 3. NAVETTE
    # -------------------------
    shuttles = db.query(Shuttle).all()

    # -------------------------
    # 4. OCCUPAZIONE PARCHEGGIO
    # -------------------------
    current = db.query(Booking).filter(
        Booking.checkin <= now,
        Booking.checkout >= now
    ).count()

    # per area
    areas = {}
    all_areas = db.query(Booking.parking_area).distinct().all()
    for (area,) in all_areas:
        if not area:
            continue
        count = db.query(Booking).filter(
            Booking.parking_area == area,
            Booking.checkin <= now,
            Booking.checkout >= now
        ).count()
        areas[area] = count

    # per fascia oraria (ogni 2 ore)
    hourly = []
    today = now.date()
    for h in range(0, 24, 2):
        start = datetime(today.year, today.month, today.day, h)
        end = start + timedelta(hours=2)

        count = db.query(Booking).filter(
            Booking.checkin <= end,
            Booking.checkout >= start
        ).count()

        hourly.append({
            "range": f"{h:02d}:00 - {h+2:02d}:00",
            "cars": count
        })

    db.close()

    return {
        "timestamp": now,
        "arrivals": arrivals,
        "departures": departures,
        "shuttles": shuttles,
        "occupancy": {
            "current": current,
            "by_area": areas,
            "hourly": hourly
        }
    }
# -------------------------
# DASHBOARD NAVETTE IN TEMPO REALE
# -------------------------

@router.get("/shuttle/dashboard")
def shuttle_dashboard():
    now = datetime.now()
    db = SessionLocal()

    # NAVETTE
    shuttles = db.query(Shuttle).all()

    # NAVETTE PER STATO
    idle = [s for s in shuttles if s.status == "idle"]
    loading = [s for s in shuttles if s.status == "loading"]
    in_transit = [s for s in shuttles if s.status == "in_transit"]
    returning = [s for s in shuttles if s.status == "returning"]

    # ARRIVI (prossime 2 ore)
    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(hours=2)
    ).order_by(Booking.checkin.asc()).all()

    # PARTENZE (prossime 2 ore)
    departures = db.query(Booking).filter(
        Booking.checkout >= now,
        Booking.checkout <= now + timedelta(hours=2)
    ).order_by(Booking.checkout.asc()).all()

    # ULTIME ASSEGNAZIONI
    assignments = db.query(ShuttleAssignment).order_by(
        ShuttleAssignment.timestamp.desc()
    ).limit(10).all()

    # PREPARA RISPOSTA
    result = {
        "timestamp": now,
        "shuttles": {
            "all": shuttles,
            "idle": idle,
            "loading": loading,
            "in_transit": in_transit,
            "returning": returning,
        },
        "arrivals": arrivals,
        "departures": departures,
        "recent_assignments": [
            {
                "shuttle": a.shuttle_id,
                "booking": a.booking_id,
                "timestamp": a.timestamp
            }
            for a in assignments
        ]
    }

    db.close()
    return result
# -------------------------
# SISTEMA DI NOTIFICHE OPERATIVE
# -------------------------

@router.get("/notifications")
def notifications():
    now = datetime.now()
    db = SessionLocal()

    notifications = []

    # -------------------------
    # 1. ARRIVI IMMINENTI (entro 30 min)
    # -------------------------
    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(minutes=30)
    ).order_by(Booking.checkin.asc()).all()

    for b in arrivals:
        notifications.append({
            "type": "arrival_imminent",
            "message": f"Arrivo imminente: {b.customer_name} ({b.car_plate})",
            "time": b.checkin,
            "booking_id": b.id
        })

    # -------------------------
    # 2. PARTENZE IMMINENTI (entro 30 min)
    # -------------------------
    departures = db.query(Booking).filter(
        Booking.checkout >= now,
        Booking.checkout <= now + timedelta(minutes=30)
    ).order_by(Booking.checkout.asc()).all()

    for b in departures:
        notifications.append({
            "type": "departure_imminent",
            "message": f"Partenza imminente: {b.customer_name} ({b.car_plate})",
            "time": b.checkout,
            "booking_id": b.id
        })

    # -------------------------
    # 3. NAVETTE DISPONIBILI
    # -------------------------
    idle_shuttles = db.query(Shuttle).filter(Shuttle.status == "idle").all()

    if idle_shuttles:
        notifications.append({
            "type": "shuttle_available",
            "message": f"{len(idle_shuttles)} navette disponibili",
            "shuttles": [s.name for s in idle_shuttles]
        })

    # -------------------------
    # 4. NAVETTE IN RITARDO (loading da troppo tempo)
    # -------------------------
    loading_shuttles = db.query(Shuttle).filter(Shuttle.status == "loading").all()

    for shuttle in loading_shuttles:
        # trova l'ultima assegnazione
        assignment = db.query(ShuttleAssignment).filter(
            ShuttleAssignment.shuttle_id == shuttle.id
        ).order_by(ShuttleAssignment.timestamp.desc()).first()

        if assignment and (now - assignment.timestamp).seconds > 900:  # 15 minuti
            notifications.append({
                "type": "shuttle_delay",
                "message": f"La navetta {shuttle.name} è in ritardo",
                "shuttle": shuttle.name,
                "since": assignment.timestamp
            })

    # -------------------------
    # 5. PRENOTAZIONI CRITICHE (molti passeggeri)
    # -------------------------
    critical = db.query(Booking).filter(
        Booking.passenger_count >= 5,
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(hours=2)
    ).all()

    for b in critical:
        notifications.append({
            "type": "critical_booking",
            "message": f"Prenotazione critica: {b.passenger_count} passeggeri per {b.customer_name}",
            "booking_id": b.id,
            "time": b.checkin
        })

    # -------------------------
    # 6. ALERT: nessuna navetta disponibile
    # -------------------------
    total_shuttles = db.query(Shuttle).count()
    busy_shuttles = db.query(Shuttle).filter(Shuttle.status != "idle").count()

    if total_shuttles > 0 and busy_shuttles == total_shuttles:
        notifications.append({
            "type": "no_shuttles_available",
            "message": "Tutte le navette sono occupate"
        })

    db.close()

    # Ordina le notifiche per importanza
    notifications = sorted(notifications, key=lambda x: x.get("time", now))

    return {
        "timestamp": now,
        "count": len(notifications),
        "notifications": notifications
    }
# -------------------------
# DASHBOARD AMMINISTRATIVA
# -------------------------

@router.get("/admin/dashboard")
def admin_dashboard():
    now = datetime.now()
    today = now.date()
    db = SessionLocal()

    # -------------------------
    # 1. RICAVI TOTALI
    # -------------------------
    bookings = db.query(Booking).all()

    def parse_amount(a):
        if not a:
            return 0
        try:
            return float(str(a).replace("€", "").replace(",", ".").strip())
        except:
            return 0

    total_revenue = sum(parse_amount(b.amount) for b in bookings)

    # -------------------------
    # 2. RICAVI PER PORTALE
    # -------------------------
    revenue_by_portal = {}
    for b in bookings:
        portal = b.portal or "unknown"
        revenue_by_portal.setdefault(portal, 0)
        revenue_by_portal[portal] += parse_amount(b.amount)

    # -------------------------
    # 3. PRENOTAZIONI PER PORTALE
    # -------------------------
    bookings_by_portal = {}
    for b in bookings:
        portal = b.portal or "unknown"
        bookings_by_portal.setdefault(portal, 0)
        bookings_by_portal[portal] += 1

    # -------------------------
    # 4. DURATA MEDIA SOGGIORNO
    # -------------------------
    durations = []
    for b in bookings:
        if b.checkin and b.checkout:
            durations.append((b.checkout - b.checkin).days)

    avg_duration = sum(durations) / len(durations) if durations else 0

    # -------------------------
    # 5. PASSEGGERI TOTALI
    # -------------------------
    total_passengers = sum(b.passenger_count or 0 for b in bookings)

    # -------------------------
    # 6. TASSO OCCUPAZIONE MEDIO (OGGI)
    # -------------------------
    start_day = datetime(today.year, today.month, today.day)
    end_day = start_day + timedelta(days=1)

    occupancy_today = db.query(Booking).filter(
        Booking.checkin <= end_day,
        Booking.checkout >= start_day
    ).count()

    # -------------------------
    # 7. ANDAMENTO SETTIMANALE (7 giorni)
    # -------------------------
    weekly = []
    for i in range(7):
        day = today - timedelta(days=i)
        start = datetime(day.year, day.month, day.day)
        end = start + timedelta(days=1)

        count = db.query(Booking).filter(
            Booking.checkin <= end,
            Booking.checkout >= start
        ).count()

        weekly.append({
            "date": str(day),
            "cars": count
        })

    weekly.reverse()

    # -------------------------
    # 8. PRENOTAZIONI PER GIORNO (check-in)
    # -------------------------
    bookings_per_day = {}
    for b in bookings:
        if b.checkin:
            d = b.checkin.date().isoformat()
            bookings_per_day.setdefault(d, 0)
            bookings_per_day[d] += 1

    # -------------------------
    # 9. PRENOTAZIONI PER FASCIA ORARIA
    # -------------------------
    hourly = {f"{h:02d}:00": 0 for h in range(24)}

    for b in bookings:
        if b.checkin:
            hour = b.checkin.hour
            hourly[f"{hour:02d}:00"] += 1

    db.close()

    return {
        "timestamp": now,
        "kpi": {
            "total_revenue": total_revenue,
            "total_bookings": len(bookings),
            "total_passengers": total_passengers,
            "avg_duration_days": avg_duration,
            "occupancy_today": occupancy_today,
        },
        "revenue_by_portal": revenue_by_portal,
        "bookings_by_portal": bookings_by_portal,
        "weekly_occupancy": weekly,
        "bookings_per_day": bookings_per_day,
        "bookings_per_hour": hourly
    }
# -------------------------
# DASHBOARD FINANZIARIA
# -------------------------

@router.get("/admin/finance")
def admin_finance():
    now = datetime.now()
    db = SessionLocal()

    bookings = db.query(Booking).all()

    # Funzione per convertire importi
    def parse_amount(a):
        if not a:
            return 0
        try:
            return float(str(a).replace("€", "").replace(",", ".").strip())
        except:
            return 0

    # -------------------------
    # 1. RICAVI PER MESE (ultimi 12 mesi)
    # -------------------------
    monthly_revenue = {}
    monthly_bookings = {}

    for b in bookings:
        if not b.checkin:
            continue
        key = b.checkin.strftime("%Y-%m")
        monthly_revenue.setdefault(key, 0)
        monthly_bookings.setdefault(key, 0)
        monthly_revenue[key] += parse_amount(b.amount)
        monthly_bookings[key] += 1

    # Ordina per data
    monthly_revenue = dict(sorted(monthly_revenue.items()))
    monthly_bookings = dict(sorted(monthly_bookings.items()))

    # -------------------------
    # 2. RICAVI PER PORTALE PER MESE
    # -------------------------
    revenue_by_portal_month = {}

    for b in bookings:
        if not b.checkin:
            continue
        month = b.checkin.strftime("%Y-%m")
        portal = b.portal or "unknown"

        revenue_by_portal_month.setdefault(portal, {})
        revenue_by_portal_month[portal].setdefault(month, 0)
        revenue_by_portal_month[portal][month] += parse_amount(b.amount)

    # -------------------------
    # 3. RICAVO MEDIO PER PRENOTAZIONE
    # -------------------------
    total_revenue = sum(parse_amount(b.amount) for b in bookings)
    avg_revenue = total_revenue / len(bookings) if bookings else 0

    # -------------------------
    # 4. FORECAST (prossimi 3 mesi)
    # Metodo semplice: media ultimi 3 mesi
    # -------------------------
    last_3_months = list(monthly_revenue.values())[-3:] if len(monthly_revenue) >= 3 else list(monthly_revenue.values())
    avg_last_3 = sum(last_3_months) / len(last_3_months) if last_3_months else 0

    forecast = {
        "next_month": avg_last_3,
        "in_2_months": avg_last_3 * 1.02,  # +2% crescita
        "in_3_months": avg_last_3 * 1.05,  # +5% crescita
    }

    # -------------------------
    # 5. STAGIONALITÀ (media ricavi per mese dell’anno)
    # -------------------------
    seasonality = {}
    for b in bookings:
        if not b.checkin:
            continue
        month = b.checkin.strftime("%m")
        seasonality.setdefault(month, [])
        seasonality[month].append(parse_amount(b.amount))

    seasonality = {m: (sum(v) / len(v)) for m, v in seasonality.items()}

    # -------------------------
    # 6. RICAVI PER GIORNO DELLA SETTIMANA
    # -------------------------
    weekday_revenue = {i: 0 for i in range(7)}  # 0 = lunedì

    for b in bookings:
        if b.checkin:
            weekday = b.checkin.weekday()
            weekday_revenue[weekday] += parse_amount(b.amount)

    # -------------------------
    # 7. RICAVI PER FASCIA ORARIA
    # -------------------------
    hourly_revenue = {f"{h:02d}:00": 0 for h in range(24)}

    for b in bookings:
        if b.checkin:
            hour = b.checkin.hour
            hourly_revenue[f"{hour:02d}:00"] += parse_amount(b.amount)

    db.close()

    return {
        "timestamp": now,
        "monthly_revenue": monthly_revenue,
        "monthly_bookings": monthly_bookings,
        "revenue_by_portal_month": revenue_by_portal_month,
        "avg_revenue_per_booking": avg_revenue,
        "forecast": forecast,
        "seasonality": seasonality,
        "weekday_revenue": weekday_revenue,
        "hourly_revenue": hourly_revenue
    }
# -------------------------
# DASHBOARD OPERATORI
# -------------------------

@router.get("/operators/dashboard")
def operators_dashboard():
    now = datetime.now()
    today = now.date()
    db = SessionLocal()

    # -------------------------
    # 1. ARRIVI E PARTENZE (prossime 2 ore)
    # -------------------------
    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(hours=2)
    ).order_by(Booking.checkin.asc()).all()

    departures = db.query(Booking).filter(
        Booking.checkout >= now,
        Booking.checkout <= now + timedelta(hours=2)
    ).order_by(Booking.checkout.asc()).all()

    # -------------------------
    # 2. PASSEGGERI TOTALI (prossime 2 ore)
    # -------------------------
    passengers_arrivals = sum(b.passenger_count or 0 for b in arrivals)
    passengers_departures = sum(b.passenger_count or 0 for b in departures)

    # -------------------------
    # 3. PRENOTAZIONI CRITICHE
    # -------------------------
    critical = [
        b for b in arrivals
        if (b.passenger_count or 0) >= 5
    ]

    # -------------------------
    # 4. STATO NAVETTE
    # -------------------------
    shuttles = db.query(Shuttle).all()

    idle = [s for s in shuttles if s.status == "idle"]
    loading = [s for s in shuttles if s.status == "loading"]
    in_transit = [s for s in shuttles if s.status == "in_transit"]
    returning = [s for s in shuttles if s.status == "returning"]

    # -------------------------
    # 5. PERFORMANCE NAVETTE (oggi)
    # -------------------------
    start_day = datetime(today.year, today.month, today.day)

    assignments_today = db.query(ShuttleAssignment).filter(
        ShuttleAssignment.timestamp >= start_day
    ).all()

    trips_by_shuttle = {}
    for a in assignments_today:
        trips_by_shuttle.setdefault(a.shuttle_id, 0)
        trips_by_shuttle[a.shuttle_id] += 1

    # -------------------------
    # 6. TEMPO MEDIO DI CARICO E VIAGGIO (stima)
    # -------------------------
    avg_loading_time = 7  # minuti (stima operativa)
    avg_trip_time = 12    # minuti (stima operativa)

    # -------------------------
    # 7. ALERT OPERATIVI
    # -------------------------
    alerts = []

    # arrivi simultanei
    if len(arrivals) >= 5:
        alerts.append("Molti arrivi simultanei nelle prossime 2 ore")

    # navette insufficienti
    if len(idle) == 0 and len(arrivals) > 0:
        alerts.append("Nessuna navetta disponibile per gli arrivi imminenti")

    # operatori sovraccarichi (stima)
    if passengers_arrivals + passengers_departures > 40:
        alerts.append("Carico operatori molto elevato")

    db.close()

    return {
        "timestamp": now,
        "workload": {
            "arrivals": len(arrivals),
            "departures": len(departures),
            "passengers_arrivals": passengers_arrivals,
            "passengers_departures": passengers_departures,
            "critical_bookings": len(critical),
        },
        "shuttles": {
            "total": len(shuttles),
            "idle": len(idle),
            "loading": len(loading),
            "in_transit": len(in_transit),
            "returning": len(returning),
            "trips_today": trips_by_shuttle,
        },
        "performance": {
            "avg_loading_time_min": avg_loading_time,
            "avg_trip_time_min": avg_trip_time,
        },
        "alerts": alerts
    }
# -------------------------
# DASHBOARD PORTALI
# -------------------------

@router.get("/portals/dashboard")
def portals_dashboard():
    now = datetime.now()
    db = SessionLocal()

    bookings = db.query(Booking).all()

    # Funzione per convertire importi
    def parse_amount(a):
        if not a:
            return 0
        try:
            return float(str(a).replace("€", "").replace(",", ".").strip())
        except:
            return 0

    # -------------------------
    # 1. DATI PER PORTALE
    # -------------------------
    portals = {}

    for b in bookings:
        portal = b.portal or "unknown"
        portals.setdefault(portal, {
            "bookings": 0,
            "revenue": 0,
            "passengers": 0,
            "durations": [],
            "critical": 0,
            "cancelled": 0,
        })

        portals[portal]["bookings"] += 1
        portals[portal]["revenue"] += parse_amount(b.amount)
        portals[portal]["passengers"] += b.passenger_count or 0

        if b.checkin and b.checkout:
            portals[portal]["durations"].append((b.checkout - b.checkin).days)

        if (b.passenger_count or 0) >= 5:
            portals[portal]["critical"] += 1

        if b.status and "cancel" in b.status.lower():
            portals[portal]["cancelled"] += 1

    # Calcoli finali
    for portal, data in portals.items():
        data["avg_revenue"] = data["revenue"] / data["bookings"] if data["bookings"] else 0
        data["avg_passengers"] = data["passengers"] / data["bookings"] if data["bookings"] else 0
        data["avg_duration"] = (
            sum(data["durations"]) / len(data["durations"])
            if data["durations"] else 0
        )
        data["cancel_rate"] = (
            data["cancelled"] / data["bookings"] * 100
            if data["bookings"] else 0
        )

    # -------------------------
    # 2. RANKING PORTALI
    # -------------------------
    ranking_revenue = sorted(portals.items(), key=lambda x: x[1]["revenue"], reverse=True)
    ranking_bookings = sorted(portals.items(), key=lambda x: x[1]["bookings"], reverse=True)
    ranking_duration = sorted(portals.items(), key=lambda x: x[1]["avg_duration"], reverse=True)
    ranking_passengers = sorted(portals.items(), key=lambda x: x[1]["avg_passengers"], reverse=True)

    # -------------------------
    # 3. ANDAMENTO MENSILE PER PORTALE
    # -------------------------
    monthly = {}

    for b in bookings:
        if not b.checkin:
            continue
        month = b.checkin.strftime("%Y-%m")
        portal = b.portal or "unknown"

        monthly.setdefault(portal, {})
        monthly[portal].setdefault(month, {
            "bookings": 0,
            "revenue": 0
        })

        monthly[portal][month]["bookings"] += 1
        monthly[portal][month]["revenue"] += parse_amount(b.amount)

    db.close()

    return {
        "timestamp": now,
        "portals": portals,
        "ranking": {
            "by_revenue": ranking_revenue,
            "by_bookings": ranking_bookings,
            "by_duration": ranking_duration,
            "by_passengers": ranking_passengers,
        },
        "monthly": monthly
    }
# -------------------------
# ASSEGNAZIONE NAVETTE AVANZATA (AI-BASED)
# -------------------------

@router.post("/shuttle/assign/optimal")
def assign_optimal_shuttle():
    now = datetime.now()
    db = SessionLocal()

    # 1. PRENOTAZIONI IMMINENTI (entro 45 minuti)
    upcoming = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(minutes=45)
    ).order_by(Booking.checkin.asc()).all()

    if not upcoming:
        db.close()
        return {"assigned": False, "reason": "Nessuna prenotazione imminente"}

    # Ordina per criticità (più passeggeri = più urgente)
    upcoming.sort(key=lambda b: (-(b.passenger_count or 1), b.checkin))

    target_booking = upcoming[0]

    # 2. NAVETTE DISPONIBILI
    idle_shuttles = db.query(Shuttle).filter(Shuttle.status == "idle").all()

    # Filtra per capacità sufficiente
    idle_shuttles = [
        s for s in idle_shuttles
        if s.capacity >= (target_booking.passenger_count or 1)
    ]

    # Bilanciamento: navette meno utilizzate oggi
    start_day = datetime(now.year, now.month, now.day)
    assignments_today = db.query(ShuttleAssignment).filter(
        ShuttleAssignment.timestamp >= start_day
    ).all()

    usage = {}
    for a in assignments_today:
        usage.setdefault(a.shuttle_id, 0)
        usage[a.shuttle_id] += 1

    def shuttle_usage(s):
        return usage.get(s.id, 0)

    idle_shuttles.sort(key=lambda s: shuttle_usage(s))

    # 3. SE C'È UNA NAVETTA IDLE → ASSEGNA
    if idle_shuttles:
        chosen = idle_shuttles[0]

        assignment = ShuttleAssignment(
            shuttle_id=chosen.id,
            booking_id=target_booking.id,
            timestamp=now
        )

        chosen.status = "loading"

        db.add(assignment)
        db.commit()
        db.refresh(assignment)
        db.close()

        return {
            "assigned": True,
            "strategy": "idle_shuttle",
            "shuttle": chosen.name,
            "booking_id": target_booking.id,
            "customer": target_booking.customer_name,
            "passengers": target_booking.passenger_count,
            "checkin": target_booking.checkin
        }

    # 4. SE NESSUNA NAVETTA È IDLE → CERCA QUELLA CHE RIENTRA PRIMA
    returning = db.query(Shuttle).filter(Shuttle.status == "returning").all()

    if returning:
        chosen = returning[0]  # prima navetta che rientra

        assignment = ShuttleAssignment(
            shuttle_id=chosen.id,
            booking_id=target_booking.id,
            timestamp=now
        )

        chosen.status = "loading"

        db.add(assignment)
        db.commit()
        db.refresh(assignment)
        db.close()

        return {
            "assigned": True,
            "strategy": "returning_shuttle",
            "shuttle": chosen.name,
            "booking_id": target_booking.id,
            "customer": target_booking.customer_name,
            "passengers": target_booking.passenger_count,
            "checkin": target_booking.checkin
        }

    # 5. SE NESSUNA NAVETTA È DISPONIBILE → ALERT
    db.close()
    return {
        "assigned": False,
        "reason": "Nessuna navetta disponibile",
        "booking_id": target_booking.id,
        "customer": target_booking.customer_name,
        "passengers": target_booking.passenger_count,
        "checkin": target_booking.checkin
    }
# -------------------------
# SISTEMA TURNI OPERATORI
# -------------------------

@router.post("/operators/create")
def create_operator(name: str, role: str):
    db = SessionLocal()
    op = Operator(name=name, role=role)
    db.add(op)
    db.commit()
    db.refresh(op)
    db.close()
    return {"created": op}


@router.post("/operators/shift/create")
def create_shift(operator_id: int, start: datetime, end: datetime):
    db = SessionLocal()

    op = db.query(Operator).filter(Operator.id == operator_id).first()
    if not op:
        db.close()
        raise HTTPException(status_code=404, detail="Operatore non trovato")

    shift = OperatorShift(operator_id=operator_id, start=start, end=end)
    db.add(shift)
    db.commit()
    db.refresh(shift)
    db.close()

    return {"created": shift}


@router.get("/operators/active")
def active_operators():
    now = datetime.now()
    db = SessionLocal()

    active = db.query(Operator).join(OperatorShift).filter(
        OperatorShift.start <= now,
        OperatorShift.end >= now
    ).all()

    db.close()
    return active


@router.get("/operators/dashboard")
def operators_shift_dashboard():
    now = datetime.now()
    db = SessionLocal()

    # operatori attivi
    active = db.query(Operator).join(OperatorShift).filter(
        OperatorShift.start <= now,
        OperatorShift.end >= now
    ).all()

    # carico di lavoro stimato
    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(hours=2)
    ).count()

    departures = db.query(Booking).filter(
        Booking.checkout >= now,
        Booking.checkout <= now + timedelta(hours=2)
    ).count()

    workload = arrivals + departures

    # operatori sovraccarichi (stima)
    overloaded = []
    if active:
        load_per_operator = workload / len(active)
        if load_per_operator > 10:
            overloaded = [op.name for op in active]

    db.close()

    return {
        "timestamp": now,
        "active_operators": [op.name for op in active],
        "workload": workload,
        "overloaded": overloaded,
        "arrivals_next_2h": arrivals,
        "departures_next_2h": departures
    }
# -------------------------
# DASHBOARD OCCUPAZIONE AVANZATA (PREVISIONI + SIMULAZIONI)
# -------------------------

@router.get("/occupancy/advanced2")
def occupancy_advanced2():
    now = datetime.now()
    today = now.date()
    db = SessionLocal()

    bookings = db.query(Booking).all()

    # -------------------------
    # 1. OCCUPAZIONE ATTUALE
    # -------------------------
    current = db.query(Booking).filter(
        Booking.checkin <= now,
        Booking.checkout >= now
    ).count()

    # -------------------------
    # 2. PREVISIONE 24 ORE
    # -------------------------
    next_24h = []
    for hour in range(24):
        t = now + timedelta(hours=hour)
        count = db.query(Booking).filter(
            Booking.checkin <= t,
            Booking.checkout >= t
        ).count()
        next_24h.append({"time": t.isoformat(), "cars": count})

    # -------------------------
    # 3. PREVISIONE 7 GIORNI
    # -------------------------
    next_7_days = []
    for d in range(7):
        day = today + timedelta(days=d)
        start = datetime(day.year, day.month, day.day)
        end = start + timedelta(days=1)

        count = db.query(Booking).filter(
            Booking.checkin <= end,
            Booking.checkout >= start
        ).count()

        next_7_days.append({
            "date": str(day),
            "cars": count
        })

    # -------------------------
    # 4. SATURAZIONE PREVISTA
    # -------------------------
    MAX_CAPACITY = 300  # puoi modificarlo

    saturation_points = [
        p for p in next_7_days if p["cars"] >= MAX_CAPACITY
    ]

    # -------------------------
    # 5. SIMULAZIONI SCENARI
    # -------------------------
    def simulate_increase(percent):
        factor = 1 + percent / 100
        return [
            {"date": p["date"], "cars": int(p["cars"] * factor)}
            for p in next_7_days
        ]

    scenario_10 = simulate_increase(10)
    scenario_20 = simulate_increase(20)
    scenario_30 = simulate_increase(30)

    # -------------------------
    # 6. OCCUPAZIONE PER GIORNO DELLA SETTIMANA
    # -------------------------
    weekday = {i: 0 for i in range(7)}

    for b in bookings:
        if b.checkin and b.checkout:
            for d in range((b.checkout - b.checkin).days + 1):
                day = b.checkin + timedelta(days=d)
                weekday[day.weekday()] += 1

    # -------------------------
    # 7. OCCUPAZIONE PER MESE
    # -------------------------
    monthly = {}

    for b in bookings:
        if b.checkin and b.checkout:
            month = b.checkin.strftime("%Y-%m")
            monthly.setdefault(month, 0)
            monthly[month] += 1

    # -------------------------
    # 8. DURATA MEDIA SOGGIORNO
    # -------------------------
    durations = [
        (b.checkout - b.checkin).days
        for b in bookings if b.checkin and b.checkout
    ]

    avg_duration = sum(durations) / len(durations) if durations else 0

    # -------------------------
    # 9. TEMPO MEDIO DI SOVRAPPOSIZIONE
    # -------------------------
    overlaps = 0
    for b in bookings:
        overlaps += db.query(Booking).filter(
            Booking.id != b.id,
            Booking.checkin <= b.checkout,
            Booking.checkout >= b.checkin
        ).count()

    avg_overlap = overlaps / len(bookings) if bookings else 0

    db.close()

    return {
        "timestamp": now,
        "current": current,
        "forecast_24h": next_24h,
        "forecast_7_days": next_7_days,
        "saturation_points": saturation_points,
        "scenarios": {
            "plus_10_percent": scenario_10,
            "plus_20_percent": scenario_20,
            "plus_30_percent": scenario_30,
        },
        "weekday_occupancy": weekday,
        "monthly_occupancy": monthly,
        "avg_duration_days": avg_duration,
        "avg_overlap": avg_overlap
    }
# -------------------------
# SISTEMA DI OTTIMIZZAZIONE NAVETTE (MULTI-ARRIVO)
# -------------------------

@router.post("/shuttle/optimize")
def optimize_shuttles():
    now = datetime.now()
    db = SessionLocal()

    # 1. ARRIVI IMMINENTI (entro 60 minuti)
    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(minutes=60)
    ).order_by(Booking.checkin.asc()).all()

    if not arrivals:
        db.close()
        return {"optimized": False, "reason": "Nessun arrivo imminente"}

    # 2. NAVETTE DISPONIBILI
    shuttles = db.query(Shuttle).filter(
        Shuttle.status.in_(["idle", "returning"])
    ).all()

    if not shuttles:
        db.close()
        return {"optimized": False, "reason": "Nessuna navetta disponibile"}

    # 3. ORDINA NAVETTE PER UTILIZZO (bilanciamento)
    start_day = datetime(now.year, now.month, now.day)
    assignments_today = db.query(ShuttleAssignment).filter(
        ShuttleAssignment.timestamp >= start_day
    ).all()

    usage = {}
    for a in assignments_today:
        usage.setdefault(a.shuttle_id, 0)
        usage[a.shuttle_id] += 1

    def shuttle_usage(s):
        return usage.get(s.id, 0)

    shuttles.sort(key=lambda s: shuttle_usage(s))

    # 4. RAGGRUPPA ARRIVI PER FASCIA ORARIA (slot da 10 minuti)
    slots = {}
    for b in arrivals:
        slot_key = b.checkin.replace(minute=(b.checkin.minute // 10) * 10, second=0, microsecond=0)
        slots.setdefault(slot_key, [])
        slots[slot_key].append(b)

    # 5. ASSEGNAZIONE OTTIMIZZATA
    plan = []
    assignments = []

    for slot_time, bookings_in_slot in slots.items():
        # ordina per criticità (più passeggeri = più urgente)
        bookings_in_slot.sort(key=lambda b: -(b.passenger_count or 1))

        for booking in bookings_in_slot:
            # trova navetta con capacità sufficiente e meno usata
            suitable = [
                s for s in shuttles
                if s.capacity >= (booking.passenger_count or 1)
            ]

            if not suitable:
                assignments.append({
                    "booking_id": booking.id,
                    "customer": booking.customer_name,
                    "status": "no_shuttle_available"
                })
                continue

            chosen = suitable[0]  # navetta meno usata

            # registra assegnazione
            assignment = ShuttleAssignment(
                shuttle_id=chosen.id,
                booking_id=booking.id,
                timestamp=now
            )
            db.add(assignment)

            # aggiorna stato navetta
            chosen.status = "loading"

            assignments.append({
                "shuttle": chosen.name,
                "booking_id": booking.id,
                "customer": booking.customer_name,
                "passengers": booking.passenger_count,
                "slot": slot_time
            })

            # aggiorna uso navetta
            usage[chosen.id] = usage.get(chosen.id, 0) + 1

            # riordina navette per bilanciamento
            shuttles.sort(key=lambda s: usage.get(s.id, 0))

    db.commit()
    db.close()

    return {
        "optimized": True,
        "timestamp": now,
        "assignments": assignments
    }
# -------------------------
# PLANNER OPERATIVO GIORNALIERO
# -------------------------

@router.get("/planner/daily")
def daily_planner():
    now = datetime.now()
    today = now.date()
    db = SessionLocal()

    # -------------------------
    # 1. COSTRUZIONE TIMELINE (slot da 30 minuti)
    # -------------------------
    timeline = []
    for h in range(24):
        for m in [0, 30]:
            slot = datetime(today.year, today.month, today.day, h, m)
            timeline.append(slot)

    # -------------------------
    # 2. ARRIVI E PARTENZE PER SLOT
    # -------------------------
    def count_in_slot(start, end):
        arrivals = db.query(Booking).filter(
            Booking.checkin >= start,
            Booking.checkin < end
        ).count()

        departures = db.query(Booking).filter(
            Booking.checkout >= start,
            Booking.checkout < end
        ).count()

        passengers_arrivals = sum(
            b.passenger_count or 0
            for b in db.query(Booking).filter(
                Booking.checkin >= start,
                Booking.checkin < end
            ).all()
        )

        passengers_departures = sum(
            b.passenger_count or 0
            for b in db.query(Booking).filter(
                Booking.checkout >= start,
                Booking.checkout < end
            ).all()
        )

        return arrivals, departures, passengers_arrivals, passengers_departures

    # -------------------------
    # 3. NAVETTE DISPONIBILI
    # -------------------------
    shuttles = db.query(Shuttle).all()
    shuttle_count = len(shuttles)

    # -------------------------
    # 4. OPERATORI ATTIVI
    # -------------------------
    active_ops = db.query(Operator).join(OperatorShift).filter(
        OperatorShift.start <= now,
        OperatorShift.end >= now
    ).all()

    # -------------------------
    # 5. COSTRUZIONE PIANO SLOT PER SLOT
    # -------------------------
    plan = []

    for i, slot in enumerate(timeline):
        start = slot
        end = slot + timedelta(minutes=30)

        arrivals, departures, pax_arr, pax_dep = count_in_slot(start, end)

        # navette richieste (stima: 8 pax per navetta)
        required_shuttles = (pax_arr + pax_dep) // 8
        if (pax_arr + pax_dep) % 8 != 0:
            required_shuttles += 1

        # operatori richiesti (stima: 1 operatore ogni 10 pax)
        required_ops = (pax_arr + pax_dep) // 10
        if (pax_arr + pax_dep) % 10 != 0:
            required_ops += 1

        plan.append({
            "slot": start.strftime("%H:%M"),
            "arrivals": arrivals,
            "departures": departures,
            "passengers_arrivals": pax_arr,
            "passengers_departures": pax_dep,
            "required_shuttles": required_shuttles,
            "required_operators": required_ops,
            "shuttles_available": shuttle_count,
            "operators_available": len(active_ops),
            "shuttle_gap": shuttle_count - required_shuttles,
            "operator_gap": len(active_ops) - required_ops
        })

    # -------------------------
    # 6. ALERT OPERATIVI
    # -------------------------
    alerts = []

    for p in plan:
        if p["shuttle_gap"] < 0:
            alerts.append(f"Slot {p['slot']}: navette insufficienti")
        if p["operator_gap"] < 0:
            alerts.append(f"Slot {p['slot']}: operatori insufficienti")
        if p["passengers_arrivals"] + p["passengers_departures"] > 40:
            alerts.append(f"Slot {p['slot']}: carico molto elevato")

    # -------------------------
    # 7. KPI GIORNALIERI
    # -------------------------
    total_arrivals = sum(p["arrivals"] for p in plan)
    total_departures = sum(p["departures"] for p in plan)
    total_passengers = sum(p["passengers_arrivals"] + p["passengers_departures"] for p in plan)

    db.close()

    return {
        "timestamp": now,
        "date": str(today),
        "plan": plan,
        "alerts": alerts,
        "kpi": {
            "total_arrivals": total_arrivals,
            "total_departures": total_departures,
            "total_passengers": total_passengers,
            "max_shuttles_needed": max(p["required_shuttles"] for p in plan),
            "max_operators_needed": max(p["required_operators"] for p in plan),
        }
    }
# -------------------------
# SISTEMA DI TARIFFE DINAMICHE
# -------------------------

@router.post("/pricing/dynamic")
def dynamic_pricing(
    checkin: datetime,
    checkout: datetime,
    portal: str = "generic"
):
    now = datetime.now()
    db = SessionLocal()

    # durata soggiorno
    days = (checkout - checkin).days
    if days <= 0:
        days = 1

    # prezzo base (puoi modificarlo)
    base_price_per_day = 6.0
    base_price = base_price_per_day * days

    # -------------------------
    # 1. OCCUPAZIONE ATTUALE
    # -------------------------
    current_occ = db.query(Booking).filter(
        Booking.checkin <= now,
        Booking.checkout >= now
    ).count()

    total_capacity = 300  # modificabile
    current_rate = current_occ / total_capacity

    occ_multiplier = 1.0
    if current_rate < 0.40:
        occ_multiplier = 0.90
    elif current_rate < 0.70:
        occ_multiplier = 1.00
    elif current_rate < 0.90:
        occ_multiplier = 1.15
    else:
        occ_multiplier = 1.30

    # -------------------------
    # 2. OCCUPAZIONE PREVISTA (prossimi 3 giorni)
    # -------------------------
    forecast_multiplier = 1.0
    for d in range(3):
        day = now + timedelta(days=d)
        start = datetime(day.year, day.month, day.day)
        end = start + timedelta(days=1)

        occ = db.query(Booking).filter(
            Booking.checkin <= end,
            Booking.checkout >= start
        ).count()

        if occ / total_capacity > 0.80:
            forecast_multiplier = 1.10
            break

    # -------------------------
    # 3. STAGIONALITÀ
    # -------------------------
    month = checkin.month
    season_multiplier = 1.0

    if month in [7, 8]:  # alta stagione
        season_multiplier = 1.20
    elif month in [6, 9, 12]:  # media stagione
        season_multiplier = 1.10

    # -------------------------
    # 4. ANTICIPO PRENOTAZIONE
    # -------------------------
    days_in_advance = (checkin - now).days

    advance_multiplier = 1.0
    if days_in_advance < 2:
        advance_multiplier = 1.10  # last minute
    elif days_in_advance > 20:
        advance_multiplier = 0.95  # early booking

    # -------------------------
    # 5. DURATA SOGGIORNO
    # -------------------------
    duration_multiplier = 1.0
    if days >= 10:
        duration_multiplier = 0.90
    elif days <= 2:
        duration_multiplier = 1.10

    # -------------------------
    # 6. PORTALE
    # -------------------------
    portal_multiplier = 1.0
    if portal.lower() == "parkos":
        portal_multiplier = 1.05
    elif portal.lower() == "myparking":
        portal_multiplier = 1.00
    elif portal.lower() == "parkingmycar":
        portal_multiplier = 0.95

    # -------------------------
    # PREZZO FINALE
    # -------------------------
    final_price = base_price
    final_price *= occ_multiplier
    final_price *= forecast_multiplier
    final_price *= season_multiplier
    final_price *= advance_multiplier
    final_price *= duration_multiplier
    final_price *= portal_multiplier

    db.close()

    return {
        "timestamp": now,
        "base_price": base_price,
        "final_price": round(final_price, 2),
        "days": days,
        "multipliers": {
            "occupancy": occ_multiplier,
            "forecast": forecast_multiplier,
            "season": season_multiplier,
            "advance": advance_multiplier,
            "duration": duration_multiplier,
            "portal": portal_multiplier
        }
    }
# -------------------------
# PREVISIONE ARRIVI (AI-BASED)
# -------------------------

@router.get("/forecast/arrivals")
def forecast_arrivals():
    now = datetime.now()
    today = now.date()
    db = SessionLocal()

    bookings = db.query(Booking).all()

    # -------------------------
    # 1. COSTRUZIONE SERIE STORICA (arrivi per giorno)
    # -------------------------
    history = {}

    for b in bookings:
        if b.checkin:
            d = b.checkin.date().isoformat()
            history.setdefault(d, 0)
            history[d] += 1

    # ordina per data
    history = dict(sorted(history.items()))

    # -------------------------
    # 2. SMOOTHING ESPONENZIALE (previsione base)
    # -------------------------
    alpha = 0.3
    forecast_value = 0
    for day, value in history.items():
        forecast_value = alpha * value + (1 - alpha) * forecast_value

    # -------------------------
    # 3. PREVISIONE 7 GIORNI
    # -------------------------
    forecast_7_days = []
    for i in range(7):
        day = today + timedelta(days=i)
        forecast_7_days.append({
            "date": str(day),
            "expected_arrivals": int(forecast_value)
        })
        # aggiorna previsione (trend leggero)
        forecast_value *= 1.02

    # -------------------------
    # 4. PREVISIONE 24 ORE (fasce orarie)
    # -------------------------
    hourly_forecast = []
    base_hourly = sum(history.values()) / len(history) / 24 if history else 1

    for h in range(24):
        hourly_forecast.append({
            "hour": f"{h:02d}:00",
            "expected_arrivals": int(base_hourly * (1.1 if 7 <= h <= 11 else 1.0))
        })

    # -------------------------
    # 5. PREVISIONE 30 GIORNI (trend mensile)
    # -------------------------
    forecast_30_days = []
    monthly_growth = 1.03  # crescita stimata

    monthly_value = sum(history.values()) / len(history) if history else 10

    for i in range(30):
        day = today + timedelta(days=i)
        forecast_30_days.append({
            "date": str(day),
            "expected_arrivals": int(monthly_value)
        })
        monthly_value *= monthly_growth

    db.close()

    return {
        "timestamp": now,
        "forecast_7_days": forecast_7_days,
        "forecast_24h": hourly_forecast,
        "forecast_30_days": forecast_30_days,
        "historical_days": len(history),
        "historical_average": sum(history.values()) / len(history) if history else 0
    }
# -------------------------
# OTTIMIZZAZIONE COMPLETA (NAVETTE + OPERATORI + OCCUPAZIONE)
# -------------------------

@router.get("/optimizer/full")
def full_optimizer():
    now = datetime.now()
    today = now.date()
    db = SessionLocal()

    # -------------------------
    # DATI DI BASE
    # -------------------------
    shuttles = db.query(Shuttle).all()
    operators = db.query(Operator).join(OperatorShift).filter(
        OperatorShift.start <= now,
        OperatorShift.end >= now
    ).all()

    arrivals = db.query(Booking).filter(
        Booking.checkin >= now,
        Booking.checkin <= now + timedelta(hours=3)
    ).order_by(Booking.checkin.asc()).all()

    departures = db.query(Booking).filter(
        Booking.checkout >= now,
        Booking.checkout <= now + timedelta(hours=3)
    ).order_by(Booking.checkout.asc()).all()

    # -------------------------
    # 1. CALCOLO CARICO
    # -------------------------
    total_pax = sum((b.passenger_count or 0) for b in arrivals + departures)

    required_shuttles = max(1, total_pax // 8)
    required_operators = max(1, total_pax // 10)

    # -------------------------
    # 2. ASSEGNAZIONE NAVETTE
    # -------------------------
    shuttle_plan = []
    shuttle_index = 0

    for b in arrivals:
        if not shuttles:
            break
        chosen = shuttles[shuttle_index % len(shuttles)]
        shuttle_plan.append({
            "shuttle": chosen.name,
            "booking_id": b.id,
            "customer": b.customer_name,
            "passengers": b.passenger_count,
            "checkin": b.checkin
        })
        shuttle_index += 1

    # -------------------------
    # 3. ASSEGNAZIONE OPERATORI
    # -------------------------
    operator_plan = []
    op_index = 0

    for b in arrivals:
        if not operators:
            break
        chosen = operators[op_index % len(operators)]
        operator_plan.append({
            "operator": chosen.name,
            "booking_id": b.id,
            "customer": b.customer_name,
            "passengers": b.passenger_count,
            "checkin": b.checkin
        })
        op_index += 1

    # -------------------------
    # 4. ALERT OPERATIVI
    # -------------------------
    alerts = []

    if len(shuttles) < required_shuttles:
        alerts.append("Navette insufficienti per il carico previsto")

    if len(operators) < required_operators:
        alerts.append("Operatori insufficienti per il carico previsto")

    if total_pax > 50:
        alerts.append("Carico molto elevato nelle prossime 3 ore")

    db.close()

    return {
        "timestamp": now,
        "summary": {
            "arrivals_next_3h": len(arrivals),
            "departures_next_3h": len(departures),
            "total_passengers": total_pax,
            "required_shuttles": required_shuttles,
            "required_operators": required_operators
        },
        "shuttle_plan": shuttle_plan,
        "operator_plan": operator_plan,
        "alerts": alerts
    }
